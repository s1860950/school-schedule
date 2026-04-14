"""
Утилиты для обработки расписания и генерации Excel-файлов.
"""
from io import BytesIO
import re
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DAYS_ORDER = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
DEFAULT_CLASSES = ['1 класс', '2 класс', '3 класс', '4 класс']
TIME_PATTERN = re.compile(r'^\s*\d{1,2}:\d{2}\s*[-–—]\s*\d{1,2}:\d{2}\s*$')
CLASS_PATTERN = re.compile(r'(\d+)\s*класс', re.IGNORECASE)


def _normalize_time_value(time_str: str) -> str:
    return re.sub(r'\s*[-–—]\s*', '-', time_str.strip())


def _time_sort_key(time_str: str) -> Tuple[int, int]:
    normalized = _normalize_time_value(time_str)
    match = re.match(r'^(\d{1,2}):(\d{2})', normalized)
    if not match:
        return (99, 99)
    return (int(match.group(1)), int(match.group(2)))


def _split_markdown_cells(line: str) -> List[str]:
    stripped = line.strip()
    cells = [cell.strip() for cell in stripped.split('|')]

    if stripped.startswith('|'):
        cells = cells[1:]

    if stripped.endswith('|'):
        cells = cells[:-1]

    return [cell for cell in cells if cell]


def _is_markdown_separator(line: str) -> bool:
    cleaned = line.replace(' ', '').replace('|', '').replace(':', '')
    return bool(cleaned) and set(cleaned) == {'-'}


def parse_schedule_text(text: str) -> List[List[str]]:
    """
    Парсит текст расписания в формат таблицы.
    Пытается извлечь таблицу из текста.
    """
    lines = text.strip().split('\n')
    rows: List[List[str]] = []

    for line in lines:
        cleaned = line.strip()
        if not cleaned:
            continue

        if '|' in cleaned:
            cells = _split_markdown_cells(cleaned)
            if cells and not _is_markdown_separator(cleaned):
                rows.append(cells)
        elif '  ' in cleaned:
            cells = [cell.strip() for cell in re.split(r'\s{2,}', cleaned) if cell.strip()]
            if len(cells) >= 2:
                rows.append(cells)
        else:
            rows.append([cleaned])

    return rows


def parse_schedule_structure(text: str) -> Tuple[Dict[str, Dict[str, Dict[str, str]]], List[str], List[str]]:
    """
    Парсит расписание из markdown-таблиц в структурированный формат.
    Возвращает: (schedule_data, days_list, classes_list)

    schedule_data структура: {day: {time: {class: subject}}}
    """
    lines = text.replace('\r', '').split('\n')
    schedule_data: Dict[str, Dict[str, Dict[str, str]]] = {}
    classes_set = set()
    days_found: List[str] = []

    current_day = None
    headers: List[str] = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        day_found = None
        for day in DAYS_ORDER:
            normalized_line = line.lower().lstrip('#').strip()
            if normalized_line == day.lower() or (normalized_line.startswith(day.lower()) and '|' not in line):
                day_found = day
                current_day = day
                schedule_data.setdefault(current_day, {})
                if day not in days_found:
                    days_found.append(day)
                headers = []
                break

        if day_found:
            continue

        if '|' not in line or not current_day or _is_markdown_separator(line):
            continue

        cells = _split_markdown_cells(line)
        if not cells:
            continue

        first_cell = _normalize_time_value(cells[0])

        if not headers and (
            first_cell.lower().startswith('время')
            or TIME_PATTERN.match(first_cell)
            or any(CLASS_PATTERN.search(cell) for cell in cells)
        ):
            headers = [cells[0], *cells[1:]]
            for cell in headers[1:]:
                class_match = CLASS_PATTERN.search(cell)
                if class_match:
                    classes_set.add(f"{class_match.group(1)} класс")
            continue

        if headers and TIME_PATTERN.match(first_cell):
            schedule_data[current_day].setdefault(first_cell, {})
            for index, header in enumerate(headers[1:], start=1):
                if index >= len(cells):
                    continue
                class_match = CLASS_PATTERN.search(header)
                if not class_match:
                    continue
                subject = cells[index].strip()
                if subject:
                    schedule_data[current_day][first_cell][f"{class_match.group(1)} класс"] = subject

    classes_list = sorted(classes_set, key=lambda value: int(re.search(r'\d+', value).group())) if classes_set else DEFAULT_CLASSES
    return schedule_data, days_found if days_found else DAYS_ORDER[:2], classes_list


def render_schedule_markdown(
    schedule_data: Dict[str, Dict[str, Dict[str, str]]],
    days_list: List[str],
    classes_list: List[str],
) -> str:
    """Собирает markdown-таблицы обратно из структурированного расписания."""
    final_classes = classes_list or DEFAULT_CLASSES
    rendered_days: List[str] = []

    for day in days_list:
        if day not in schedule_data or not schedule_data[day]:
            continue

        lines = [day, ""]
        lines.append("| Время | " + " | ".join(final_classes) + " |")
        lines.append("|" + "---|" * (len(final_classes) + 1))

        for time_str in sorted(schedule_data[day].keys(), key=_time_sort_key):
            subjects = [schedule_data[day][time_str].get(class_name, '') for class_name in final_classes]
            lines.append(f"| {_normalize_time_value(time_str)} | " + " | ".join(subjects) + " |")

        rendered_days.append("\n".join(lines))

    return "\n\n".join(rendered_days)


def enforce_monday_class_hour(schedule_text: str) -> str:
    """
    Гарантирует, что в понедельник первым уроком у каждого класса стоит разговоры о важном.
    Если расписание не удалось распарсить, возвращает исходный текст.
    """
    schedule_data, days_list, classes_list = parse_schedule_structure(schedule_text)
    monday = 'Понедельник'

    if monday not in schedule_data or not schedule_data[monday]:
        return schedule_text

    monday_times = sorted(schedule_data[monday].keys(), key=_time_sort_key)
    if not monday_times:
        return schedule_text

    first_time = monday_times[0]
    monday_row = schedule_data[monday].setdefault(first_time, {})

    final_classes = classes_list or DEFAULT_CLASSES
    for class_name in final_classes:
        monday_row[class_name] = 'Разговоры о важном'

    final_days = [day for day in DAYS_ORDER if day in schedule_data and schedule_data[day]]
    if not final_days:
        final_days = days_list

    return render_schedule_markdown(schedule_data, final_days, final_classes)


def create_schedule_excel(schedule_text: str) -> BytesIO:
    """
    Создает Excel-файл с расписанием из текста.
    Возвращает BytesIO объект с содержимым файла.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Расписание"

    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    day_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    day_font = Font(bold=True, color="FFFFFF", size=11)
    time_fill = PatternFill(start_color="F8F9FF", end_color="F8F9FF", fill_type="solid")
    time_font = Font(bold=True, color="667EEA", size=10)

    border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    schedule_data, days_list, classes_list = parse_schedule_structure(schedule_text)

    if not schedule_data or not days_list:
        headers = ["Время", *DEFAULT_CLASSES]
        times = ["8:30-9:15", "9:30-10:15", "10:30-11:15", "11:30-12:15"]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        for row, time in enumerate(times, 2):
            cell = worksheet.cell(row=row, column=1)
            cell.value = time
            cell.border = border
            cell.alignment = center_alignment
            cell.font = time_font
            cell.fill = time_fill

        for row in range(2, len(times) + 2):
            for col in range(2, len(headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = left_alignment

        worksheet.column_dimensions['A'].width = 14
        for col in range(2, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 18

        worksheet.row_dimensions[1].height = 25
        for row in range(2, len(times) + 2):
            worksheet.row_dimensions[row].height = 25

    else:
        current_row = 1

        for day in days_list:
            if day not in schedule_data or not schedule_data[day]:
                continue

            day_cell = worksheet.cell(row=current_row, column=1)
            day_cell.value = day
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.border = border
            day_cell.alignment = center_alignment
            worksheet.merge_cells(f'A{current_row}:{get_column_letter(len(classes_list) + 1)}{current_row}')
            worksheet.row_dimensions[current_row].height = 25
            current_row += 1

            headers = ['Время'] + classes_list
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment

            worksheet.row_dimensions[current_row].height = 22
            current_row += 1

            times = sorted(schedule_data[day].keys(), key=_time_sort_key)
            for time_str in times:
                cell = worksheet.cell(row=current_row, column=1)
                cell.value = _normalize_time_value(time_str)
                cell.fill = time_fill
                cell.font = time_font
                cell.border = border
                cell.alignment = center_alignment

                for col, class_name in enumerate(classes_list, 2):
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.value = schedule_data[day][time_str].get(class_name, '')
                    cell.border = border
                    cell.alignment = left_alignment

                worksheet.row_dimensions[current_row].height = 20
                current_row += 1

            current_row += 1

        worksheet.column_dimensions['A'].width = 14
        for col in range(2, len(classes_list) + 2):
            worksheet.column_dimensions[get_column_letter(col)].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
